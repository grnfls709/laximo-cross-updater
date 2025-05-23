#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Скрипт для автоматизации дополнения кросс-номеров в Excel-файле номенклатуры
с использованием API Laximo.
"""

import os
import sys
import time
import logging
import xml.etree.ElementTree as ET
import requests
import openpyxl
from datetime import datetime

# Настройка логирования
log_filename = f"laximo_cross_update_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Константы
LAXIMO_BASE_URL = "https://ws.laximo.net/ws/"
EXCEL_FILE_PATH = "/home/ubuntu/upload/Номенклатура.xlsx"
OUTPUT_FILE_PATH = "/home/ubuntu/workspace/Номенклатура_с_кроссами.xlsx"
SAVE_INTERVAL = 100  # Сохранять каждые N строк
REQUEST_DELAY = 0.5  # Задержка между запросами в секундах

class LaximoAPI:
    """Класс для работы с API Laximo"""
    
    def __init__(self):
        self.session = requests.Session()
    
    def find_oem(self, oem):
        """
        Поиск детали по артикулу (OEM)
        
        Args:
            oem (str): Артикул детали
            
        Returns:
            list: Список найденных деталей с их detailid
        """
        params = {
            "FindOEM": f"Locale=ru_RU|OEM={oem}|ReplacementTypes=|Options="
        }
        
        try:
            response = self.session.get(LAXIMO_BASE_URL, params=params)
            response.raise_for_status()
            
            # Парсинг XML-ответа
            root = ET.fromstring(response.text)
            details = []
            
            for detail in root.findall(".//detail"):
                detail_info = {
                    "detailid": detail.get("detailid"),
                    "manufacturer": detail.get("manufacturer"),
                    "oem": detail.get("oem"),
                    "formattedoem": detail.get("formattedoem"),
                    "name": detail.get("name")
                }
                details.append(detail_info)
            
            return details
        
        except requests.exceptions.RequestException as e:
            logger.error(f"Ошибка при запросе FindOEM для артикула {oem}: {str(e)}")
            return []
        except ET.ParseError as e:
            logger.error(f"Ошибка при парсинге XML для артикула {oem}: {str(e)}")
            return []
    
    def find_replacements(self, detail_id):
        """
        Поиск кросс-номеров по detailid
        
        Args:
            detail_id (str): Уникальный идентификатор детали
            
        Returns:
            list: Список найденных кросс-номеров
        """
        params = {
            "FindReplacements": f"Locale=ru_RU|DetailId={detail_id}"
        }
        
        try:
            response = self.session.get(LAXIMO_BASE_URL, params=params)
            response.raise_for_status()
            
            # Парсинг XML-ответа
            root = ET.fromstring(response.text)
            replacements = []
            
            for row in root.findall(".//row"):
                replacement_info = {
                    "manufacturer": row.get("manufacturer"),
                    "oem": row.get("oem"),
                    "formattedoem": row.get("formattedoem"),
                    "name": row.get("name"),
                    "rate": row.get("rate")
                }
                replacements.append(replacement_info)
            
            return replacements
        
        except requests.exceptions.RequestException as e:
            logger.error(f"Ошибка при запросе FindReplacements для detailid {detail_id}: {str(e)}")
            return []
        except ET.ParseError as e:
            logger.error(f"Ошибка при парсинге XML для detailid {detail_id}: {str(e)}")
            return []


class ExcelProcessor:
    """Класс для обработки Excel-файла"""
    
    def __init__(self, input_file, output_file):
        """
        Инициализация обработчика Excel
        
        Args:
            input_file (str): Путь к исходному Excel-файлу
            output_file (str): Путь для сохранения результата
        """
        self.input_file = input_file
        self.output_file = output_file
        self.workbook = None
        self.sheet = None
        self.total_rows = 0
        self.processed_rows = 0
        self.success_count = 0
        self.error_count = 0
        self.no_data_count = 0
    
    def load_workbook(self):
        """Загрузка Excel-файла"""
        try:
            logger.info(f"Загрузка файла: {self.input_file}")
            self.workbook = openpyxl.load_workbook(self.input_file)
            self.sheet = self.workbook.active
            self.total_rows = self.sheet.max_row
            logger.info(f"Файл загружен. Всего строк: {self.total_rows}")
            return True
        except Exception as e:
            logger.error(f"Ошибка при загрузке Excel-файла: {str(e)}")
            return False
    
    def save_workbook(self, is_final=False):
        """
        Сохранение Excel-файла
        
        Args:
            is_final (bool): Флаг финального сохранения
        """
        try:
            save_path = self.output_file
            if not is_final:
                # Для промежуточных сохранений добавляем временную метку
                base, ext = os.path.splitext(self.output_file)
                save_path = f"{base}_temp{ext}"
            
            self.workbook.save(save_path)
            logger.info(f"Файл сохранен: {save_path}")
            return True
        except Exception as e:
            logger.error(f"Ошибка при сохранении Excel-файла: {str(e)}")
            return False
    
    def process_file(self, laximo_api):
        """
        Обработка Excel-файла и дополнение кросс-номеров
        
        Args:
            laximo_api (LaximoAPI): Экземпляр класса для работы с API Laximo
            
        Returns:
            bool: Результат обработки
        """
        if not self.load_workbook():
            return False
        
        start_time = time.time()
        
        # Начинаем со второй строки (пропускаем заголовки)
        for row_idx in range(2, self.total_rows + 1):
            try:
                # Получаем артикул из 4-го столбца
                article = self.sheet.cell(row=row_idx, column=4).value
                
                if not article:
                    logger.warning(f"Строка {row_idx}: Пустой артикул, пропускаем")
                    self.no_data_count += 1
                    continue
                
                # Очищаем артикул от лишних пробелов
                article = str(article).strip()
                
                logger.info(f"Обработка строки {row_idx}/{self.total_rows}, артикул: {article}")
                
                # Получаем текущее текстовое описание
                current_description = self.sheet.cell(row=row_idx, column=6).value or ""
                
                # Шаг 1: Получаем detailid по артикулу
                details = laximo_api.find_oem(article)
                
                if not details:
                    logger.warning(f"Строка {row_idx}: Не найдены детали для артикула {article}")
                    self.no_data_count += 1
                    continue
                
                # Шаг 2: Получаем кросс-номера для каждого найденного detailid
                all_cross_numbers = set()  # Используем множество для исключения дубликатов
                
                for detail in details:
                    detail_id = detail.get("detailid")
                    if not detail_id:
                        continue
                    
                    # Добавляем задержку между запросами
                    time.sleep(REQUEST_DELAY)
                    
                    replacements = laximo_api.find_replacements(detail_id)
                    
                    for replacement in replacements:
                        formatted_oem = replacement.get("formattedoem")
                        if formatted_oem:
                            all_cross_numbers.add(formatted_oem)
                
                if not all_cross_numbers:
                    logger.warning(f"Строка {row_idx}: Не найдены кросс-номера для артикула {article}")
                    self.no_data_count += 1
                    continue
                
                # Шаг 3: Формируем новое текстовое описание
                cross_numbers_text = ", ".join(sorted(all_cross_numbers))
                
                # Проверяем, есть ли уже кросс-номера в текущем описании
                if current_description:
                    # Если текущее описание не заканчивается на пробел или запятую, добавляем разделитель
                    if not current_description.rstrip().endswith((",", " ")):
                        current_description += ", "
                    
                    # Добавляем новые кросс-номера к существующему описанию
                    new_description = current_description + cross_numbers_text
                else:
                    new_description = cross_numbers_text
                
                # Шаг 4: Записываем новое описание в Excel
                self.sheet.cell(row=row_idx, column=6).value = new_description
                
                self.success_count += 1
                logger.info(f"Строка {row_idx}: Успешно добавлены кросс-номера: {cross_numbers_text}")
                
            except Exception as e:
                logger.error(f"Строка {row_idx}: Ошибка при обработке: {str(e)}")
                self.error_count += 1
            
            self.processed_rows += 1
            
            # Периодическое сохранение
            if self.processed_rows % SAVE_INTERVAL == 0:
                elapsed_time = time.time() - start_time
                avg_time_per_row = elapsed_time / self.processed_rows
                estimated_time_left = avg_time_per_row * (self.total_rows - self.processed_rows)
                
                logger.info(f"Прогресс: {self.processed_rows}/{self.total_rows} строк "
                           f"({self.processed_rows/self.total_rows*100:.1f}%)")
                logger.info(f"Примерное оставшееся время: {estimated_time_left/60:.1f} минут")
                
                self.save_workbook(is_final=False)
        
        # Финальное сохранение
        result = self.save_workbook(is_final=True)
        
        # Итоговая статистика
        total_time = time.time() - start_time
        logger.info("=" * 50)
        logger.info("Обработка завершена")
        logger.info(f"Всего строк: {self.total_rows}")
        logger.info(f"Обработано строк: {self.processed_rows}")
        logger.info(f"Успешно: {self.success_count}")
        logger.info(f"Ошибок: {self.error_count}")
        logger.info(f"Без данных: {self.no_data_count}")
        logger.info(f"Общее время выполнения: {total_time/60:.1f} минут")
        logger.info("=" * 50)
        
        return result


def main():
    """Основная функция"""
    logger.info("Запуск скрипта обновления кросс-номеров")
    
    # Инициализация API Laximo
    laximo_api = LaximoAPI()
    
    # Инициализация обработчика Excel
    excel_processor = ExcelProcessor(EXCEL_FILE_PATH, OUTPUT_FILE_PATH)
    
    # Обработка файла
    result = excel_processor.process_file(laximo_api)
    
    if result:
        logger.info(f"Скрипт успешно завершен. Результат сохранен в {OUTPUT_FILE_PATH}")
    else:
        logger.error("Скрипт завершен с ошибками")


if __name__ == "__main__":
    main()
