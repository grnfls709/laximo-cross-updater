#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Упрощенный тестовый скрипт для проверки функциональности обновления кросс-номеров
на минимальном наборе данных из Excel-файла.
"""

import os
import sys
import logging
import openpyxl
from datetime import datetime

# Импортируем классы из основного скрипта
sys.path.append('/home/ubuntu/workspace')
from laximo_cross_updater import LaximoAPI

# Настройка логирования
log_filename = f"mini_test_laximo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Константы
EXCEL_FILE_PATH = "/home/ubuntu/upload/Номенклатура.xlsx"
TEST_ROWS = 3  # Минимальное количество строк для тестирования


def test_api_with_sample_articles():
    """
    Тестирование API Laximo на нескольких артикулах из Excel
    
    Returns:
        bool: Результат тестирования API
    """
    try:
        # Создаем экземпляр API
        api = LaximoAPI()
        
        # Открываем Excel для получения тестовых артикулов
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = wb.active
        
        logger.info(f"Тестирование API на {TEST_ROWS} артикулах")
        
        # Проверяем первые TEST_ROWS строк
        success_count = 0
        
        for row in range(2, 2 + TEST_ROWS):
            if row > sheet.max_row:
                break
                
            test_article = sheet.cell(row=row, column=4).value
            
            if not test_article:
                logger.warning(f"Строка {row}: Пустой артикул, пропускаем")
                continue
            
            logger.info(f"Тестирование артикула: {test_article}")
            
            # Шаг 1: Поиск детали по артикулу
            details = api.find_oem(test_article)
            
            if not details:
                logger.warning(f"Строка {row}: Не найдены детали для артикула {test_article}")
                continue
            
            logger.info(f"Найдено деталей: {len(details)}")
            for i, detail in enumerate(details[:2]):  # Показываем только первые 2 для краткости
                logger.info(f"Деталь {i+1}: {detail}")
            
            if len(details) > 2:
                logger.info(f"... и еще {len(details) - 2} деталей")
            
            # Шаг 2: Поиск кросс-номеров для первой найденной детали
            detail_id = details[0].get("detailid")
            
            if not detail_id:
                logger.warning(f"Строка {row}: Не удалось получить detailid")
                continue
            
            logger.info(f"Поиск кросс-номеров для detailid: {detail_id}")
            replacements = api.find_replacements(detail_id)
            
            if not replacements:
                logger.warning(f"Строка {row}: Не найдены кросс-номера для detailid {detail_id}")
                continue
            
            logger.info(f"Найдено кросс-номеров: {len(replacements)}")
            for i, replacement in enumerate(replacements[:3]):  # Показываем только первые 3 для краткости
                logger.info(f"Кросс {i+1}: {replacement.get('formattedoem', 'Н/Д')} ({replacement.get('manufacturer', 'Н/Д')})")
            
            if len(replacements) > 3:
                logger.info(f"... и еще {len(replacements) - 3} кросс-номеров")
            
            # Формируем пример текста для добавления
            cross_numbers = [r.get("formattedoem") for r in replacements if r.get("formattedoem")]
            cross_text = ", ".join(cross_numbers[:5])  # Берем только первые 5 для примера
            
            if len(cross_numbers) > 5:
                cross_text += f"... и еще {len(cross_numbers) - 5}"
                
            logger.info(f"Пример текста для добавления: {cross_text}")
            
            success_count += 1
            logger.info(f"Тест для артикула {test_article} успешно завершен")
            logger.info("-" * 40)
        
        logger.info(f"Успешно протестировано артикулов: {success_count} из {TEST_ROWS}")
        return success_count > 0
    
    except Exception as e:
        logger.error(f"Ошибка при тестировании API: {str(e)}")
        return False


def main():
    """Основная функция тестирования"""
    logger.info("=" * 50)
    logger.info("Запуск упрощенного тестирования API Laximo")
    logger.info("=" * 50)
    
    # Тестирование API на нескольких артикулах
    if test_api_with_sample_articles():
        logger.info("=" * 50)
        logger.info("Тестирование API успешно завершено")
        logger.info("=" * 50)
    else:
        logger.error("=" * 50)
        logger.error("Тестирование API завершилось с ошибками")
        logger.error("=" * 50)


if __name__ == "__main__":
    main()
